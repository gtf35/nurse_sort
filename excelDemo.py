#!/usr/bin/env python3
# -*- coding: utf-8 -*-

' excle 模板处理模块 '

__author__ = 'gtf35'
#excle 处理框架
from openpyxl import *
import openpyxl
#配置文件
import config


def createDemo():
    workBook = Workbook()
    nurseWorkSheet = createWorkSheet(workBook, config.nurseWorkSheetName)
    settingsWorkSheet = createWorkSheet(workBook, config.settingsWorkSheetName)
    outputWorkSheet = createWorkSheet(workBook, config.outputWorkSheetName)
    nurseWorkSheet[config.nurseWorkSheetTipCellAddr] = config.nurseWorkSheetTipCellText
    settingsWorkSheet[config.settingsSheetTipCellAddr] = config.settingsSheetTipCellText
    settingsWorkSheet[config.settingsSheetBeginYearCellAddr] = config.settingsSheetBeginYearCellText
    settingsWorkSheet[config.settingsSheetEndYearCellAddr] = config.settingsSheetEndYearCellText
    settingsWorkSheet[config.settingsSheetBeginMonthCellAddr] = config.settingsSheetBeginMonthCellText
    settingsWorkSheet[config.settingsSheetEndMonthCellAddr] = config.settingsSheetEndMonthCellText
    #删除默认的那个
    workBook.remove(workBook["Sheet"])
    workBook.save(config.demoExcelPath)

def createWorkSheet(workbook, title, color = config.workSheetDefaultColor):
    workSheet = workbook.create_sheet()
    workSheet.title = title
    workSheet.sheet_properties.tabColor = color
    return workSheet

def checkIsDemo(path):
    try:
        workbook = openpyxl.load_workbook(path)
        nurseWorkSheet = workbook.worksheets[0]
        settingsWorkSheet = workbook.worksheets[1]
        if nurseWorkSheet[config.nurseWorkSheetTipCellAddr].value != config.nurseWorkSheetTipCellText:
            return False
        if settingsWorkSheet[config.settingsSheetTipCellAddr].value != config.settingsSheetTipCellText:
            return False
        if settingsWorkSheet[config.settingsSheetBeginYearCellAddr].value != config.settingsSheetBeginYearCellText:
            return False
        if settingsWorkSheet[config.settingsSheetEndYearCellAddr].value != config.settingsSheetEndYearCellText:
            return False
        if settingsWorkSheet[config.settingsSheetBeginMonthCellAddr].value != config.settingsSheetBeginMonthCellText:
            return False
        if settingsWorkSheet[config.settingsSheetEndMonthCellAddr].value != config.settingsSheetEndMonthCellText:
            return False
        return True
    except BaseException as e:
        print(e)
        return False

if __name__ == '__main__':
    #print(checkIsDemo(config.demoExcelPath))
    createDemo()